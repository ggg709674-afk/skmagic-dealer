# -*- coding: utf-8 -*-
"""SK매직 수수료표 엑셀 -> web/assets/commission.js (window.COMMISSION_DB) 재생성.

admin.js parseCommissionWorkbook 과 동일한 파싱 규칙:
- 컬럼(0-index): B=1 품목, C=2 모델, D=3 코드, E=4 컬러/형태, F=5 의무,
  H=7 관리주기, I=8 기준가, J=9 운영가/기본할인, K=10 전사할인, L=11 타사보상
- 13행(idx 12)부터 데이터, 병합 셀은 좌상단 값 전파
- 기본요금 = 전사할인(>0) 우선, 없으면 운영가
- 홈페이지 등록 모델만(products.json, 코드 base 9자리), 색상 변형은 base 10자리로 1행
- ★ 수수료합계는 보안상 정적 파일에 절대 넣지 않음 (admin.js 상단 주석 참조)

사용: python gen_commission_js.py "<수수료표.xlsx>" "<source 표기>"
"""
import json, pathlib, re, sys, datetime
import openpyxl

ROOT = pathlib.Path(__file__).parent
PJ = ROOT / "data" / "products.json"
DST = ROOT / "web" / "assets" / "commission.js"

MAIN = {"100000005", "100000010", "100000024", "100000245", "1000000245"}
MAT_SIZE = {"S": "SS", "Q": "Q", "K": "K"}

def com_base_code(code):
    s = str(code or "")
    return s[:3] + s[4:] if re.match(r"^MAT[SQK]", s) else s

def com_size(code):
    s = str(code or "")
    m = re.match(r"^MAT([SQK])", s)
    return MAT_SIZE.get(s[3], s[3]) if (m and len(s) > 3) else ""

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return v
    s = str(v).replace(",", "").replace(" ", "")
    return float(s) if re.match(r"^-?\d+(\.\d+)?$", s) else None

def form(e, pummok, cycle):
    if "비데" in (pummok or ""):
        mo = re.sub(r"[^\d]", "", str(cycle or ""))
        mo = int(mo) if mo else 0
        if mo >= 12: return "셀프형"
        if mo > 0: return "방문형"
        return "셀프형" if re.search(r"lite", str(e), re.I) else "방문형"
    return "셀프형" if "셀프형" in str(e) else "방문형"

def main():
    xlsx = sys.argv[1]
    source = sys.argv[2] if len(sys.argv) > 2 else pathlib.Path(xlsx).stem

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    sn = next((n for n in wb.sheetnames if "수수료" in n), wb.sheetnames[0])
    ws = wb[sn]

    # grid[r][c] (0-index) + 병합 전파
    grid = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                grid.setdefault(cell.row - 1, {})[cell.column - 1] = cell.value
    for mg in ws.merged_cells.ranges:
        v = grid.get(mg.min_row - 1, {}).get(mg.min_col - 1)
        if v is None: continue
        for r in range(mg.min_row - 1, mg.max_row):
            for c in range(mg.min_col - 1, mg.max_col):
                grid.setdefault(r, {}).setdefault(c, v)

    g = lambda r, c: ("" if grid.get(r, {}).get(c) is None else str(grid[r][c]))

    CB, CC, CD, CE, CF, CH, CI, CJ, CK, CL = 1, 2, 3, 4, 5, 7, 8, 9, 10, 11
    rows, lastB = [], ""
    for r in range(12, ws.max_row):
        b, c, d, e, f = g(r, CB), g(r, CC), g(r, CD), g(r, CE), g(r, CF)
        h, i, j, k, l = g(r, CH), g(r, CI), g(r, CJ), g(r, CK), g(r, CL)
        if not c and not d: continue
        uimu, gijun = num(f), num(i)
        if uimu is None and gijun is None: continue
        pummok = b or lastB
        if b: lastB = b
        jeonsa, unyeong = num(k), num(j)
        gibon = jeonsa if (jeonsa is not None and jeonsa > 0) else unyeong
        pummok = pummok.replace("메트리스", "매트리스")
        rows.append({
            "품목": pummok,
            "모델": re.sub(r"\s+", " ", c).strip(),
            "코드": d,
            "사이즈": com_size(d),
            "형태": form(e, pummok, h),
            "의무": int(uimu) if uimu is not None else None,
            "관리주기": h,
            "기준가": int(gijun) if gijun is not None else None,
            "기본요금": int(gibon) if gibon is not None else None,
            "타사보상": int(num(l)) if num(l) is not None else None,
        })

    # 홈페이지 등록 모델만
    db = json.loads(PJ.read_text(encoding="utf-8"))
    home_base = set()
    for p in db["products"]:
        if not p.get("model"): continue
        if not any(cat in MAIN for cat in (p.get("categories") or [])): continue
        home_base.add(com_base_code(p["model"].split("\n")[0].strip())[:9])
    rows = [x for x in rows if com_base_code(x["코드"])[:9] in home_base]

    # 색상 묶음 dedupe — 품목|코드base10|형태|의무|사이즈
    seen, out = set(), []
    for x in rows:
        key = f"{x['품목']}|{com_base_code(x['코드'])[:10]}|{x['형태']}|{x['의무']}|{x['사이즈']}"
        if key in seen: continue
        seen.add(key)
        out.append(x)

    pummok_sun = list(dict.fromkeys(x["품목"] for x in out))
    payload = {
        "source": source,
        "built_at": datetime.date.today().isoformat(),
        "품목순": pummok_sun,
        "rows": out,
    }
    js = (
        "/* auto-generated from SK매직 수수료표 — DO NOT EDIT */\n"
        "/* ★ 수수료합계(본부 원본 수수료)는 보안상 정적 파일에서 제거됨 — 산하는 get_commission_scoped RPC 로만 차감값 수신, "
        "본부는 commission_data DB 에서 원본 read. 생성기 수정해도 이 필드는 넣지 말 것. */\n"
        f"window.COMMISSION_DB = {json.dumps(payload, ensure_ascii=False)};\n"
    )
    DST.write_text(js, encoding="utf-8")
    print(f"[done] {len(out)} rows ({', '.join(pummok_sun)}) -> {DST}")
    mini = [x for x in out if "WPUIAC606" in x["코드"]]
    print(f"[check] WPUIAC606 rows: {len(mini)}")
    for x in mini: print("  ", json.dumps(x, ensure_ascii=False))

if __name__ == "__main__":
    main()
